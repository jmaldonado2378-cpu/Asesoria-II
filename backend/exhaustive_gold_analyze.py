import openpyxl

def exhaustive_gold_analyze(path):
    print(f"\n--- INSPECCIÓN EXHAUSTIVA: {path} ---")
    try:
        wb = openpyxl.load_workbook(path, data_only=False) # Cargamos fórmulas también
        print(f"Hojas encontradas: {wb.sheetnames}")
        for sn in wb.sheetnames:
            ws = wb[sn]
            print(f"\n[Hoja: {sn}]")
            for r in range(1, 20):
                row_vals = []
                for c in range(1, 12):
                    cell = ws.cell(r, c)
                    val = cell.value if cell.value else "."
                    row_vals.append(str(val))
                if any(v != "." for v in row_vals):
                    print(f"R{r}: {row_vals}")
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    exhaustive_gold_analyze(r"c:\Users\Usuario\Documents\App asesor\backend\lab\static\templates\reclamo.xlsx")
