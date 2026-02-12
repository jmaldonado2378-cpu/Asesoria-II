import openpyxl
import os
import glob

def find_by_string(directory, target_string):
    files = glob.glob(os.path.join(directory, "*.xlsx"))
    for f in files:
        try:
            wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
            for sn in wb.sheetnames:
                ws = wb[sn]
                for r in range(1, 40):
                    for c in range(1, 15):
                        val = str(ws.cell(r, c).value)
                        if target_string.lower() in val.lower():
                            print(f"MATCH: {f} | Sheet: {sn} | Cell: ({r},{c}) | Value: {val}")
                            return f
        except:
            continue
    return None

if __name__ == "__main__":
    print("Buscando 'Cliente Directo'...")
    match = find_by_string(r"C:\Users\Usuario\Downloads", "Cliente Directo")
    if not match:
        print("No se encontró 'Cliente Directo'.")
    
    print("\nBuscando 'Harinas y Panificados'...")
    match2 = find_by_string(r"C:\Users\Usuario\Downloads", "Harinas y Panificados")
    if not match2:
        print("No se encontró 'Harinas y Panificados'.")
