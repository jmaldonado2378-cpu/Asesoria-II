import openpyxl

def find_labels_in_file(path):
    print(f"\n--- BUSCANDO LABELS EN: {path} ---")
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        for sn in wb.sheetnames:
            ws = wb[sn]
            print(f"Hoja: {sn}")
            for r in range(1, 40):
                row_vals = [str(ws.cell(r, c).value) if ws.cell(r, c).value else "" for c in range(1, 15)]
                row_str = " | ".join(row_vals)
                if "cliente directo" in row_str.lower() or "contacto" in row_str.lower():
                    print(f"ENCONTRADO en R{r}: {row_str}")
                if "harinas y panificados" in row_str.lower():
                    print(f"ENCONTRADO en R{r}: {row_str}")
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    find_labels_in_file(r"C:\Users\Usuario\Downloads\la definitiva.xlsx")
    find_labels_in_file(r"C:\Users\Usuario\Downloads\ultima.xlsx")
