import openpyxl

def exhaustive_analyze(path):
    print(f"\n--- {path} ---")
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        print(f"Hoja: {ws.title}")
        for r in range(1, 50):
            row = [str(ws.cell(r, c).value) if ws.cell(r, c).value else "" for c in range(1, 10)]
            if any(row):
                print(f"R{r}: {row}")
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    exhaustive_analyze(r"C:\Users\Usuario\Downloads\tucu.xlsx")
