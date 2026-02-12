import openpyxl

def final_template_check(path):
    print(f"\n--- VALIDACIÓN FINAL DEL TEMPLATE GOOGLE: {path} ---")
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        print(f"Hojas: {wb.sheetnames}")
        ws = wb.active
        print(f"Hoja activa: {ws.title}")
        
        # Verificamos celdas clave solicitadas por el usuario
        # Fila 1: GESTIÓN TÉCNICA Y DESARROLLO (C1:G3)
        # Fila 4: Harinas y Panificados (C4:G5)
        # Fila 6: Mapping dinámico (B6, F6)
        # Fila 9: Labels de Reclamo (Cliente Directo, etc)
        
        checks = {
            "C1 (Header)": ws['C1'].value,
            "C4 (Slogan)": ws['C4'].value,
            "A9 (Label R9)": ws['A9'].value,
            "B6 (Cliente Placehold)": ws['B6'].value,
            "F6 (Proyecto Placehold)": ws['F6'].value
        }
        
        for label, val in checks.items():
            print(f"{label}: '{val}'")
            
        # Inspeccionamos fila 9 completa
        r9_vals = [str(ws.cell(9, c).value) if ws.cell(9, c).value else "" for c in range(1, 15)]
        print(f"Contenido R9: {r9_vals}")

    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    final_template_check(r"C:\Users\Usuario\Downloads\la definitiva.xlsx")
