import openpyxl
import os
import glob

def find_template(directory):
    files = glob.glob(os.path.join(directory, "*.xlsx"))
    for f in files:
        if "Reclamo_oficial" in f: continue
        try:
            wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
            ws = wb.active
            # Buscamos en C1 o celdas cercanas
            val_c1 = str(ws['C1'].value) if ws['C1'].value else ""
            val_c2 = str(ws['C2'].value) if ws['C2'].value else ""
            val_c3 = str(ws['C3'].value) if ws['C3'].value else ""
            
            content = val_c1 + val_c2 + val_c3
            if "GESTIÓN TÉCNICA" in content or "DESARROLLO" in content:
                print(f"MATCH ENCONTRADO: {f}")
                print(f"C1-C3: {content}")
                print(f"C4: {ws['C4'].value}")
                return f
        except:
            continue
    return None

if __name__ == "__main__":
    match = find_template(r"C:\Users\Usuario\Downloads")
    if not match:
        print("No se encontró el template en Downloads.")
    else:
        print(f"Template identificado: {match}")
