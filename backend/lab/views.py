import io
import os
import xlsxwriter
try:
    from xhtml2pdf import pisa
    from django.template.loader import render_to_string
except ImportError:
    pisa = None
    render_to_string = None

from django.utils import timezone
from django.http import HttpResponse, FileResponse
from django.shortcuts import get_object_or_404
from rest_framework import viewsets, status
from rest_framework.decorators import api_view, parser_classes
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from datetime import datetime
from django.conf import settings
from .models import (
    Client, Project, Ensayo, Ingredient, 
    ProjectIngredientPrice, Visit, EnsayoDetail, EnsayoImage,
    TechnicalReport, Complaint, ComplaintImage
)
import base64
import urllib.request
import urllib.error
from .serializers import (
    ClientSerializer, 
    ProjectSerializer, 
    EnsayoSerializer, 
    IngredientSerializer,
    ProjectIngredientPriceSerializer, 
    VisitSerializer,
    EnsayoDetailSerializer,
    EnsayoImageSerializer,
    TechnicalReportSerializer,
    ComplaintSerializer,
    ComplaintImageSerializer
)

# --- UTILIDADES GLOBALES (REPORTES) ---

def resolve_image_path(path):
    """
    Resuelve una ruta de imagen de forma robusta, buscando en MEDIA_ROOT
    si la ruta absoluta falla. Útil para migraciones entre entornos (Win/Linux).
    """
    if not path: return None
    from django.conf import settings
    
    # 1. Intentar ruta absoluta directa
    if os.path.exists(path):
        return path
        
    # 2. Intentar buscar el nombre del archivo dentro de MEDIA_ROOT
    filename = os.path.basename(path)
    final_path = os.path.join(settings.MEDIA_ROOT, filename)
    if os.path.exists(final_path):
        return final_path
        
    # 3. Intentar reconstruir desde la ruta relativa si contiene 'media'
    try:
        if 'media' in path:
            relative_path = path.split('media')[-1].lstrip('\\').lstrip('/')
            final_path = os.path.join(settings.MEDIA_ROOT, relative_path)
            if os.path.exists(final_path):
                return final_path
    except:
        pass
        
    # 4. Búsqueda profunda (último recurso)
    for root, dirs, files in os.walk(settings.MEDIA_ROOT):
        if filename in files:
            return os.path.join(root, filename)
            
    return None

def parse_smart_date(val):
    """Parsea fechas en múltiples formatos (ISO, DD/MM/YYYY, etc)"""
    if not val: return None
    if isinstance(val, (datetime, timezone.datetime)):
        return val.date()
    val_str = str(val).split('T')[0] # Limpiar ISO strings con tiempo
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(val_str, fmt).date()
        except ValueError:
            continue
    return None

def get_image_base64(path):
    """
    Lee un archivo de imagen y lo devuelve como string Base64.
    """
    resolved_path = resolve_image_path(path)
    if not resolved_path: return None
    try:
        with open(resolved_path, "rb") as image_file:
            encoded_string = base64.b64encode(image_file.read()).decode('utf-8')
            ext = os.path.splitext(resolved_path)[1].lower().replace('.', '')
            if ext not in ['jpg', 'jpeg', 'png', 'gif']: ext = 'jpeg'
            return f"data:image/{ext};base64,{encoded_string}"
    except Exception as e:
        print(f"ERROR Base64: {str(e)}")
        return None
    return None

def link_callback(uri, rel):
    """
    Convierte URIs de MEDIA a rutas absolutas en disco para xhtml2pdf.
    Mantiene URLs externas intactas.
    """
    if uri.startswith('http://') or uri.startswith('https://') or uri.startswith('data:'):
        return uri
        
    from django.conf import settings
    if uri.startswith(settings.MEDIA_URL):
        path = os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ""))
    elif os.path.isabs(uri):
        path = uri
    else:
        path = os.path.join(settings.MEDIA_ROOT, uri)
        
    resolved = resolve_image_path(path)
    return resolved if resolved else uri

class ClientViewSet(viewsets.ModelViewSet):
    queryset = Client.objects.all()
    serializer_class = ClientSerializer

class ProjectViewSet(viewsets.ModelViewSet):
    queryset = Project.objects.all()
    serializer_class = ProjectSerializer

class IngredientViewSet(viewsets.ModelViewSet):
    queryset = Ingredient.objects.all()
    serializer_class = IngredientSerializer

class EnsayoViewSet(viewsets.ModelViewSet):
    queryset = Ensayo.objects.all().order_by('-date')
    serializer_class = EnsayoSerializer

class EnsayoDetailViewSet(viewsets.ModelViewSet):
    queryset = EnsayoDetail.objects.all()
    serializer_class = EnsayoDetailSerializer

class EnsayoImageViewSet(viewsets.ModelViewSet):
    queryset = EnsayoImage.objects.all()
    serializer_class = EnsayoImageSerializer

    def create(self, request, *args, **kwargs):
        if 'image' in request.FILES:
            file_obj = request.FILES['image']
            
            supabase_url = settings.SUPABASE_URL
            supabase_key = settings.SUPABASE_KEY
            if not supabase_url or not supabase_key:
                return Response({'error': 'Supabase no configurado en el backend'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
            try:
                ext = os.path.splitext(file_obj.name)[1]
                safe_name = file_obj.name.replace(' ', '_')
                filename = f"{timezone.now().strftime('%Y%m%d_%H%M%S')}_{safe_name}"
                
                bucket_name = "ensayo_photos"
                upload_url = f"{supabase_url}/storage/v1/object/{bucket_name}/{filename}"
                req_upload = urllib.request.Request(upload_url, data=file_obj.read(), method="POST")
                req_upload.add_header("Authorization", f"Bearer {supabase_key}")
                req_upload.add_header("apikey", supabase_key)
                req_upload.add_header("Content-Type", file_obj.content_type or "application/octet-stream")
                
                try:
                    with urllib.request.urlopen(req_upload, timeout=15.0) as res:
                        pass
                except urllib.error.HTTPError as e:
                    import traceback
                    print(f"Supabase upload error: {e.read().decode('utf-8', errors='ignore')}")
                    raise Exception(f"HTTPError {e.code}: {e.reason}")
                
                # Guardar solo el nombre del archivo (la propiedad full_url del modelo reconstruirá el resto)
                data = request.data.copy()
                data['image'] = filename
                
                serializer = self.get_serializer(data=data)
                serializer.is_valid(raise_exception=True)
                self.perform_create(serializer)
                headers = self.get_success_headers(serializer.data)
                return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)
                
            except Exception as e:
                import traceback
                print(traceback.format_exc())
                return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
                
        # Fallback if no file is sent, or if it's sent as a URL string directly
        return super().create(request, *args, **kwargs)

class ProjectIngredientPriceViewSet(viewsets.ModelViewSet):
    queryset = ProjectIngredientPrice.objects.all()
    serializer_class = ProjectIngredientPriceSerializer

class VisitViewSet(viewsets.ModelViewSet):
    queryset = Visit.objects.all().order_by('-date')
    serializer_class = VisitSerializer

class TechnicalReportViewSet(viewsets.ModelViewSet):
    queryset = TechnicalReport.objects.all().order_by('-report_date', '-created_at')
    serializer_class = TechnicalReportSerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        project_id = self.request.query_params.get('project')
        if project_id:
            queryset = queryset.filter(project_id=project_id)
        return queryset

class ComplaintViewSet(viewsets.ModelViewSet):
    queryset = Complaint.objects.all().order_by('-loading_date', '-created_at')
    serializer_class = ComplaintSerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        project_id = self.request.query_params.get('project')
        if project_id:
            queryset = queryset.filter(project_id=project_id)
        return queryset

class ComplaintImageViewSet(viewsets.ModelViewSet):
    queryset = ComplaintImage.objects.all()
    serializer_class = ComplaintImageSerializer

    def create(self, request, *args, **kwargs):
        if 'image' in request.FILES:
            file_obj = request.FILES['image']
            
            supabase_url = settings.SUPABASE_URL
            supabase_key = settings.SUPABASE_KEY
            if not supabase_url or not supabase_key:
                return Response({'error': 'Supabase no configurado en el backend'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
            try:
                ext = os.path.splitext(file_obj.name)[1]
                safe_name = file_obj.name.replace(' ', '_')
                filename = f"{timezone.now().strftime('%Y%m%d_%H%M%S')}_{safe_name}"
                
                # Reusing 'ensayo_photos' bucket
                bucket_name = "ensayo_photos" 
                
                upload_url = f"{supabase_url}/storage/v1/object/{bucket_name}/complaints/{filename}"
                req_upload = urllib.request.Request(upload_url, data=file_obj.read(), method="POST")
                req_upload.add_header("Authorization", f"Bearer {supabase_key}")
                req_upload.add_header("apikey", supabase_key)
                req_upload.add_header("Content-Type", file_obj.content_type or "application/octet-stream")
                
                try:
                    with urllib.request.urlopen(req_upload, timeout=15.0) as res:
                        pass
                except urllib.error.HTTPError as e:
                    import traceback
                    print(f"Supabase upload error: {e.read().decode('utf-8', errors='ignore')}")
                    raise Exception(f"HTTPError {e.code}: {e.reason}")
                
                # Guardar la ruta relativa dentro del bucket
                data = request.data.copy()
                data['image'] = f"complaints/{filename}"
                
                serializer = self.get_serializer(data=data)
                serializer.is_valid(raise_exception=True)
                self.perform_create(serializer)
                headers = self.get_success_headers(serializer.data)
                return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)
                
            except Exception as e:
                import traceback
                print(traceback.format_exc())
                return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
                
        return super().create(request, *args, **kwargs)

# --- RECLAMOS TÉCNICOS (CARGA MANUAL EXCLUSIVA) ---
# Se ha eliminado la lógica de importación masiva para garantizar la integridad de los datos manuales.

@api_view(['GET'])
def generar_reporte_reclamo_estandar(request):
    """
    GENERADOR DE REPORTE DE RECLAMO ESTÁNDAR
    Mapeo de Inyección:
    - Cliente Directo -> B6
    - Proyecto -> F6
    - Lote -> E12
    - Tipo Harina -> B14
    - Producto Final -> B16 
    - Proceso -> F16
    - Descripción -> B18
    """
    import openpyxl
    complaint_id = request.query_params.get('complaint')
    complaint = None
    if complaint_id:
        complaint = get_object_or_404(Complaint, id=complaint_id)

    template_path = os.path.join(os.path.dirname(__file__), 'static', 'templates', 'reclamo.xlsx')
    
    if not os.path.exists(template_path):
        return Response({"error": "Template file not found."}, status=404)

    wb = openpyxl.load_workbook(template_path)
    if "PLANTILLA_RECLAMOS" in wb.sheetnames:
        ws = wb["PLANTILLA_RECLAMOS"]
    else:
        ws = wb.active

    # --- INYECCIÓN DINÁMICA DE CARGA MANUAL ---
    if complaint:
        ws['B6'] = complaint.direct_client or (complaint.project.client.name if complaint.project.client else "---")
        ws['F6'] = complaint.project.name
        ws['E12'] = complaint.batch or "---"
        ws['B14'] = complaint.flour_type or "---"
        ws['B16'] = complaint.product_made or "---"
        ws['F16'] = complaint.process_type or "---"
        ws['B18'] = complaint.description or "---"
    else:
        # Fallback a proyecto si no hay ID de reclamo (descarga de plantilla limpia con datos de cliente)
        project_id = request.query_params.get('project')
        if project_id:
            project = Project.objects.filter(id=project_id).first()
            if project:
                ws['B6'] = project.client.name if project.client else "---"
                ws['F6'] = project.name

    # --- INYECCIÓN DE IMÁGENES (ANEXO) ---
    if complaint and complaint.images.exists():
        ws_photos = wb.create_sheet("ANEXO FOTOS")
        ws_photos.column_dimensions['A'].width = 80
        curr_row = 1
        
        for img in complaint.images.all():
            if img.image:
                try:
                    req_get = urllib.request.Request(str(img.full_url))
                    with urllib.request.urlopen(req_get, timeout=10.0) as response:
                        if response.status == 200:
                            from openpyxl.drawing.image import Image as OpenpyxlImage
                            
                            # Título de la foto
                            title = f"FOTO: {img.caption or 'Sin nota'}"
                            ws_photos.cell(row=curr_row, column=1, value=title)
                            ws_photos.cell(row=curr_row, column=1).font = openpyxl.styles.Font(bold=True, size=12)
                            curr_row += 1

                            # Insertar Imagen desde RAM
                            img_stream = io.BytesIO(response.read())
                            img_data = OpenpyxlImage(img_stream)
                        
                        # Redimensión proporcional para el Excel
                        orig_w, orig_h = img_data.width, img_data.height
                        aspect = orig_w / orig_h
                        img_data.width = 500
                        img_data.height = 500 / aspect
                        
                        ws_photos.add_image(img_data, f'A{curr_row}')
                        
                        # Espaciado (aproximadamente la altura de la imagen en filas)
                        rows_to_skip = int(img_data.height / 15) + 2
                        curr_row += rows_to_skip
                except Exception as e:
                    print(f"Error inyectando imagen {img.id}: {e}")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"RECLAMO_{complaint.batch or 'S_LOTE'}.xlsx" if complaint else "reporte_reclamo_estandar.xlsx"
    response = HttpResponse(
        output.read(), 
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename={filename}"
    return response


# --- VISTAS DE REPORTES ---

@api_view(['POST'])
def generar_informe_tecnico_estandar(request):

    project_id = request.data.get('project')
    start_date = request.data.get('start_date')
    end_date = request.data.get('end_date')
    conclusions = request.data.get('technical_observations', '')
    requested_format = request.data.get('format', 'excel').lower()
    save_to_history = request.data.get('save_to_history', False)

    if not project_id:
        return Response({"error": "ID de proyecto es requerido."}, status=400)

    try:
        project = get_object_or_404(Project, id=project_id)
        
        # Parseo robusto de fechas
        s_date = parse_smart_date(start_date) or datetime(2000, 1, 1).date()
        e_date = parse_smart_date(end_date) or datetime(2100, 12, 31).date()

        # Datos técnicos
        essays = Ensayo.objects.filter(project=project, date__range=[s_date, e_date]).order_by('date')
        visits = Visit.objects.filter(project=project, date__range=[s_date, e_date]).order_by('date')
        complaints = Complaint.objects.filter(project=project, loading_date__range=[s_date, e_date]).order_by('loading_date')

        # Nombre de archivo estandarizado
        client_name = str(project.client.name).replace(' ', '_') if project.client else "Sin_Cliente"
        proj_name = str(project.name).replace(' ', '_')
        report_date_str = str(request.data.get('report_date', timezone.now().strftime('%Y-%m-%d')))
        
        # Preparar datos para inyección dinámica de nombres de harina en ensayos
        essays_data = []
        for e in essays:
            base_flour_detail = e.details.filter(ingredient__is_base_flour=True).first()
            e_data = {
                'code': e.code,
                'date': e.date,
                'base_flour_name': base_flour_detail.ingredient.name if base_flour_detail else "No especificada",
                'description': e.description,
                'final_score': e.final_score,
                'conclusion': e.conclusion
            }
            essays_data.append(e_data)

        print(f"DEBUG: Generando reporte para proyecto {project.id}, save_to_history={save_to_history}")
        # --- GUARDAR EN HISTORIAL (Opcional) ---
        if save_to_history:
            try:
                TechnicalReport.objects.create(
                    project=project,
                    report_date=timezone.now().date(),
                    start_date=s_date,
                    end_date=e_date,
                    technical_observations=conclusions
                )
                print("DEBUG: Historial guardado.")
            except Exception as he:
                print(f"DEBUG ERROR: No se pudo guardar historial: {str(he)}")

        # --- GENERACIÓN DE CONTENIDO ---
        
        if requested_format == 'pdf':
            if pisa is None:
                return Response({"error": "Librería xhtml2pdf no está instalada en el servidor."}, status=500)
                
            # Logo institucional en Base64 para el PDF
            logo_b64_req = request.data.get('logo_b64')
            if logo_b64_req:
                logo_b64 = logo_b64_req
            else:
                logo_path = os.path.join(settings.BASE_DIR, 'lab', 'static', 'images', 'logo_institucional.png')
                logo_b64 = get_image_base64(logo_path)

            # Pre-serializar ensayos con imágenes para el template (xhtml2pdf no soporta ORM lazy loading)
            essays_for_pdf = []
            for e in essays:
                base_flour_detail = e.details.filter(ingredient__is_base_flour=True).first()
                score_num = float(e.final_score) if e.final_score else 0
                essays_for_pdf.append({
                    'code': e.code or f'ENS-{e.id}',
                    'date': e.date,
                    'base_flour_name': base_flour_detail.ingredient.name if base_flour_detail else "No especificada",
                    'final_score': f"{score_num:.1f}" if e.final_score else None,
                    'final_score_num': score_num,
                    'conclusion': e.conclusion or '',
                })

            # Pre-serializar imágenes de ensayos
            essays_with_imgs = []
            for e in essays:
                imgs = [{'image': img.full_url, 'caption': img.caption or ''} for img in e.images.all()]
                if imgs:
                    essays_with_imgs.append({'code': e.code or f'ENS-{e.id}', 'images': imgs})


            # Pre-serializar imágenes de reclamos
            complaints_with_imgs = []
            for c in complaints:
                imgs = [{'image': img.full_url, 'caption': img.caption or ''} for img in c.images.all()]
                if imgs:
                    complaints_with_imgs.append({'loading_date': c.loading_date, 'images': imgs})

            context = {
                'project': project,
                'start_date': s_date,
                'end_date': e_date,
                'date': timezone.now(),
                'conclusions': str(conclusions) if conclusions else "",
                'essays': essays_for_pdf,
                'visits': visits,
                'complaints': complaints,
                'logo_b64': logo_b64,
                'essays_data_with_images': essays_with_imgs,
                'complaints_with_images': complaints_with_imgs,
            }
            html = render_to_string('reports/gestion_reporte_pdf.html', context)
            buffer = io.BytesIO()
            pisa_status = pisa.CreatePDF(
                html,
                dest=buffer,
                link_callback=link_callback
            )

            
            if pisa_status.err:
                return Response({"error": "Error al generar PDF vía xhtml2pdf"}, status=400)
                
            buffer.seek(0)
            filename = f"IT_{client_name}_{proj_name}_{report_date_str}.pdf"
            return FileResponse(buffer, as_attachment=True, filename=filename)


        # --- LÓGICA EXCEL ESTÁNDAR ---
        import xlsxwriter
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output)
        ws = workbook.add_worksheet("INFORME TÉCNICO")

        # Formatos
        fmt_title = workbook.add_format({'bold': True, 'font_size': 18, 'font_color': 'white', 'bg_color': '#1A1A1A', 'align': 'center', 'valign': 'vcenter', 'border': 1})
        fmt_subtitle = workbook.add_format({'font_size': 12, 'font_color': 'white', 'bg_color': '#404040', 'align': 'center', 'valign': 'vcenter', 'border': 1})
        fmt_label = workbook.add_format({'bold': True, 'bg_color': '#F1F5F9', 'border': 2})
        fmt_value = workbook.add_format({'border': 2, 'bg_color': 'white'})
        fmt_header = workbook.add_format({'bold': True, 'font_color': 'white', 'bg_color': '#333333', 'align': 'center', 'valign': 'vcenter', 'border': 1, 'text_wrap': True})
        fmt_sec = workbook.add_format({'bold': True, 'font_color': 'white', 'bg_color': '#333333', 'border': 1})
        fmt_legal = workbook.add_format({'italic': True, 'font_size': 9, 'align': 'center'})
        fmt_ph_label = workbook.add_format({'bold': True, 'bg_color': '#333333', 'font_color': 'white', 'border': 1})

        # 1. ENCABEZADO
        ws.merge_range('A1:B5', "", workbook.add_format({'border': 2}))
        logo_path = os.path.join(os.path.dirname(__file__), 'static', 'images', 'logo_institucional.png')
        if os.path.exists(logo_path):
            ws.insert_image('A1', logo_path, {'x_offset': 10, 'y_offset': 10, 'x_scale': 0.12, 'y_scale': 0.12})

        ws.merge_range('C1:G3', "GESTIÓN TÉCNICA Y DESARROLLO", fmt_title)
        ws.merge_range('C4:G5', "Harinas y Panificados", fmt_subtitle)
        ws.merge_range('H1:I5', "REFERENCIA\n\nMOD-RECLAMO", workbook.add_format({'border': 2, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True}))

        # 2. DATOS
        ws.write('A7', 'CLIENTE:', fmt_label)
        ws.merge_range('B7:D7', project.client.name if project.client else "---", fmt_value)
        ws.write('E7', 'PROYECTO:', fmt_label)
        ws.merge_range('F7:H7', project.name, fmt_value)

        # 3. CONTENIDO
        curr = 9
        ws.merge_range(curr, 0, curr, 8, "CONCLUSIONES Y OBSERVACIONES TÉCNICAS", fmt_sec)
        curr += 1
        ws.merge_range(curr, 0, curr+4, 8, str(conclusions or ''), workbook.add_format({'border': 1, 'valign': 'top', 'text_wrap': True}))
        curr += 6

        ws.merge_range(curr, 0, curr, 8, "RESULTADOS DE ENSAYOS", fmt_sec)
        curr += 1
        h_labels = ["CÓDIGO", "FECHA", "HARINA BASE", "DESCRIPCIÓN", "PUNTAJE", "CONCLUSIÓN"]
        col_widths = [25, 20, 15, 15, 15, 20, 20, 60]
        for i, (label, w) in enumerate(zip(h_labels, col_widths)):
            ws.set_column(i, i, w)
            ws.write(curr, i, label, fmt_header)
        
        curr += 1
        for ed in essays_data:
            ws.write(curr, 0, str(ed.get('code', '')), fmt_value)
            ws.write(curr, 1, str(ed.get('date', '')), fmt_value)
            ws.write(curr, 2, str(ed.get('base_flour_name', '')), fmt_value)
            ws.write(curr, 3, str(ed.get('description', '')), fmt_value)
            ws.write(curr, 4, str(ed.get('final_score', 0)), fmt_value)
            ws.write(curr, 5, str(ed.get('conclusion', '')), fmt_value)
            curr += 1

        # 4. ANEXO FOTOGRÁFICO (Hoja Nueva)
        ws_photos = workbook.add_worksheet("ANEXO FOTOS")
        ws_photos.set_column('A:A', 80) # Ancho para la foto
        row_ph = 1
        
        def insert_safely(worksheet, row, col, img_obj, title):
            nonlocal row_ph
            worksheet.write(row, col, title, fmt_ph_label)
            row_ph += 1
            
            if img_obj and img_obj.full_url:
                try:
                    req_get = urllib.request.Request(str(img_obj.full_url))
                    with urllib.request.urlopen(req_get, timeout=10.0) as response:
                        if response.status == 200:
                            img_stream = io.BytesIO(response.read())
                            worksheet.insert_image(row_ph, col, "image.jpg", {
                                'image_data': img_stream,
                                'x_scale': 0.5, 
                                'y_scale': 0.5,
                                'object_position': 1
                            })
                            row_ph += 25
                        else:
                            worksheet.write(row_ph, col, "ERROR: No se pudo descargar la imagen", workbook.add_format({'font_color': 'red'}))
                            row_ph += 2
                except Exception as e:
                    worksheet.write(row_ph, col, f"ERROR: Insertando foto ({e})", workbook.add_format({'font_color': 'red'}))
                    row_ph += 2
            else:
                worksheet.write(row_ph, col, "SIN IMAGEN", workbook.add_format({'bg_color': '#CCCCCC'}))
                row_ph += 2

        # Inyectar fotos de Ensayos
        for essay in essays:
            for img in essay.images.all():
                insert_safely(ws_photos, row_ph, 0, img, f"ENSAYO {essay.code}")

        # Inyectar fotos de Reclamos
        for complaint in complaints:
            for img in complaint.images.all():
                insert_safely(ws_photos, row_ph, 0, img, f"RECLAMO {complaint.loading_date}")

        if row_ph == 1:
            ws_photos.write('A1', "No se encontraron imágenes para este período de auditoría.")

        workbook.close()
        buffer = io.BytesIO()
        buffer.write(output.getvalue())
        buffer.seek(0)
        filename = f"IT_{client_name}_{proj_name}_{report_date_str}.xlsx"
        return FileResponse(buffer, as_attachment=True, filename=filename)

    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print("--- CRITICAL ERROR IN REPORT GENERATION ---")
        print(error_details)
        print("-------------------------------------------")
        # EXPOMEMOS EL TRACEBACK PARA DEPURACIÓN EN DESPLIEGUE
        return Response({
            "error": f"Error en generación final: {str(e)}",
            "traceback": error_details 
        }, status=500)
    


def serve_media_view(request, path):
    """
    Vista personalizada para servir archivos media en producción (Render).
    Solo para debugging o entornos controlados donde Whitenoise no maneja media.
    """
    file_path = os.path.join(settings.MEDIA_ROOT, path)
    if os.path.exists(file_path):
        return FileResponse(open(file_path, 'rb'))
    return HttpResponse(status=404)
