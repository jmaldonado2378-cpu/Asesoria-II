import io
import os
import urllib.request
import urllib.error
from django.conf import settings
from django.utils import timezone
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from rest_framework import viewsets, status
from rest_framework.decorators import api_view
from rest_framework.response import Response
from lab.models import Complaint, ComplaintImage, Project
from lab.serializers import ComplaintSerializer, ComplaintImageSerializer

class ComplaintViewSet(viewsets.ModelViewSet):
    queryset = Complaint.objects.all().order_by('-loading_date', '-created_at')
    serializer_class = ComplaintSerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        project_id = self.request.query_params.get('project')
        if project_id:
            queryset = queryset.filter(project_id=project_id)
        return queryset

class ComplaintImageViewSet(viewsets.ModelViewSet):
    queryset = ComplaintImage.objects.all()
    serializer_class = ComplaintImageSerializer

    def create(self, request, *args, **kwargs):
        if 'image' in request.FILES:
            file_obj = request.FILES['image']
            
            supabase_url = settings.SUPABASE_URL
            supabase_key = settings.SUPABASE_KEY
            if not supabase_url or not supabase_key:
                return Response({'error': 'Supabase no configurado en el backend'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
            try:
                ext = os.path.splitext(file_obj.name)[1]
                safe_name = file_obj.name.replace(' ', '_')
                filename = f"{timezone.now().strftime('%Y%m%d_%H%M%S')}_{safe_name}"
                
                # Reusing 'ensayo_photos' bucket
                bucket_name = "ensayo_photos" 
                
                upload_url = f"{supabase_url}/storage/v1/object/{bucket_name}/complaints/{filename}"
                req_upload = urllib.request.Request(upload_url, data=file_obj.read(), method="POST")
                req_upload.add_header("Authorization", f"Bearer {supabase_key}")
                req_upload.add_header("apikey", supabase_key)
                req_upload.add_header("Content-Type", file_obj.content_type or "application/octet-stream")
                
                try:
                    with urllib.request.urlopen(req_upload, timeout=15.0) as res:
                        pass
                except urllib.error.HTTPError as e:
                    import traceback
                    print(f"Supabase upload error: {e.read().decode('utf-8', errors='ignore')}")
                    raise Exception(f"HTTPError {e.code}: {e.reason}")
                
                # Guardar la ruta relativa dentro del bucket
                data = request.data.copy()
                data['image'] = f"complaints/{filename}"
                
                serializer = self.get_serializer(data=data)
                serializer.is_valid(raise_exception=True)
                self.perform_create(serializer)
                headers = self.get_success_headers(serializer.data)
                return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)
                
            except Exception as e:
                import traceback
                print(traceback.format_exc())
                return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
                
        return super().create(request, *args, **kwargs)

# --- RECLAMOS TÉCNICOS (CARGA MANUAL EXCLUSIVA) ---
@api_view(['GET'])
def generar_reporte_reclamo_estandar(request):
    """
    GENERADOR DE REPORTE DE RECLAMO ESTÁNDAR
    Mapeo de Inyección:
    - Cliente Directo -> B6
    - Proyecto -> F6
    - Lote -> E12
    - Tipo Harina -> B14
    - Producto Final -> B16 
    - Proceso -> F16
    - Descripción -> B18
    """
    import openpyxl
    complaint_id = request.query_params.get('complaint')
    complaint = None
    if complaint_id:
        complaint = get_object_or_404(Complaint, id=complaint_id)

    # Need to properly resolve the path considering we are in `views/` now
    # Original path: os.path.join(os.path.dirname(__file__), 'static', 'templates', 'reclamo.xlsx')
    # Since we are in `views/`, we need to go up one level
    template_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static', 'templates', 'reclamo.xlsx')
    
    if not os.path.exists(template_path):
        return Response({"error": "Template file not found."}, status=404)

    wb = openpyxl.load_workbook(template_path)
    if "PLANTILLA_RECLAMOS" in wb.sheetnames:
        ws = wb["PLANTILLA_RECLAMOS"]
    else:
        ws = wb.active

    # --- INYECCIÓN DINÁMICA DE CARGA MANUAL ---
    if complaint:
        ws['B6'] = complaint.direct_client or (complaint.project.client.name if complaint.project.client else "---")
        ws['F6'] = complaint.project.name
        ws['E12'] = complaint.batch or "---"
        ws['B14'] = complaint.flour_type or "---"
        ws['B16'] = complaint.product_made or "---"
        ws['F16'] = complaint.process_type or "---"
        ws['B18'] = complaint.description or "---"
    else:
        # Fallback a proyecto si no hay ID de reclamo (descarga de plantilla limpia con datos de cliente)
        project_id = request.query_params.get('project')
        if project_id:
            project = Project.objects.filter(id=project_id).first()
            if project:
                ws['B6'] = project.client.name if project.client else "---"
                ws['F6'] = project.name

    # --- INYECCIÓN DE IMÁGENES (ANEXO) ---
    if complaint and complaint.images.exists():
        ws_photos = wb.create_sheet("ANEXO FOTOS")
        ws_photos.column_dimensions['A'].width = 80
        curr_row = 1
        
        for img in complaint.images.all():
            if img.image:
                try:
                    req_get = urllib.request.Request(str(img.full_url))
                    with urllib.request.urlopen(req_get, timeout=10.0) as response:
                        if response.status == 200:
                            from openpyxl.drawing.image import Image as OpenpyxlImage
                            
                            # Título de la foto
                            title = f"FOTO: {img.caption or 'Sin nota'}"
                            ws_photos.cell(row=curr_row, column=1, value=title)
                            ws_photos.cell(row=curr_row, column=1).font = openpyxl.styles.Font(bold=True, size=12)
                            curr_row += 1

                            # Insertar Imagen desde RAM
                            img_stream = io.BytesIO(response.read())
                            img_data = OpenpyxlImage(img_stream)
                        
                        # Redimensión proporcional para el Excel
                        orig_w, orig_h = img_data.width, img_data.height
                        aspect = orig_w / orig_h
                        img_data.width = 500
                        img_data.height = 500 / aspect
                        
                        ws_photos.add_image(img_data, f'A{curr_row}')
                        
                        # Espaciado (aproximadamente la altura de la imagen en filas)
                        rows_to_skip = int(img_data.height / 15) + 2
                        curr_row += rows_to_skip
                except Exception as e:
                    print(f"Error inyectando imagen {img.id}: {e}")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"RECLAMO_{complaint.batch or 'S_LOTE'}.xlsx" if complaint else "reporte_reclamo_estandar.xlsx"
    response = HttpResponse(
        output.read(), 
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename={filename}"
    return response
