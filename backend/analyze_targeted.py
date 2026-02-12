import openpyxl

def analyze_targeted(path):
    print(f"\n--- ANALIZANDO: {path} ---")
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        print(f"Hoja: {ws.title}")
        
        # Revisamos fila 10 para ver si es Reclamos
        r10 = [str(ws.cell(10, c).value) if ws.cell(10, c).value else "" for c in range(1, 10)]
        print(f"R10: {r10}")
        
        # Revisamos fila 6 (Mapping)
        r6 = [str(ws.cell(6, c).value) if ws.cell(6, c).value else "" for c in range(1, 10)]
        print(f"R6: {r6}")
        
        # Revisamos C1:G3
        print(f"C1: {ws['C1'].value}")
        print(f"C4: {ws['C4'].value}")
        
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    analyze_targeted(r"C:\Users\Usuario\Downloads\la definitiva.xlsx")
    analyze_targeted(r"C:\Users\Usuario\Downloads\ultima.xlsx")
