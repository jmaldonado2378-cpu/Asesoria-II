import openpyxl

def analyze_all_excels(paths):
    for path in paths:
        print(f"\n--- Analizando: {path} ---")
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                print(f"\nHoja: {sheet_name}")
                for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=10), 1):
                    row_data = [str(cell.value) if cell.value else "" for cell in row]
                    if any(row_data):
                        print(f"Fila {row_idx}: {row_data}")
        except Exception as e:
            print(f"Error cargando {path}: {e}")

if __name__ == "__main__":
    analyze_all_excels([
        r"C:\Users\Usuario\Downloads\ultima.xlsx",
        r"C:\Users\Usuario\Downloads\tucu.xlsx",
        r"C:\Users\Usuario\Downloads\lo.xlsx"
    ])
