import openpyxl

def quick_analyze(path):
    print(f"\n--- {path} ---")
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        print(f"Hojas: {wb.sheetnames}")
        for sn in wb.sheetnames:
            ws = wb[sn]
            print(f"\n[{sn}] Primeras 15 filas:")
            for r in range(1, 16):
                row = [str(ws.cell(r, c).value) if ws.cell(r, c).value else "" for c in range(1, 8)]
                if any(row):
                    print(f"R{r}: {row}")
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    quick_analyze(r"C:\Users\Usuario\Downloads\tucu.xlsx")
    quick_analyze(r"C:\Users\Usuario\Downloads\lo.xlsx")
    quick_analyze(r"C:\Users\Usuario\Downloads\ultima.xlsx")
